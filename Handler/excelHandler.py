import time
from datetime import date

from openpyxl import Workbook

from Handler.dbHandler import SQLops


class excelOps:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.filename = "./output/" + str(date.today()) + "-" + str(time.time()) + ".xlsx"

    def add(self, invoiceList):
        self.ws["A1"] = "发票号码"
        self.ws["B1"] = "发票文件名"
        self.ws["C1"] = "销售方名称"
        self.ws["D1"] = "销售方识别号"
        self.ws["E1"] = "购买方名称"
        self.ws["F1"] = "购买方识别号"
        self.ws["G1"] = "发票类型"
        self.ws["H1"] = "发票类型编码"
        self.ws["I1"] = "发票总金额"
        self.ws["J1"] = "开票日期"
        self.ws["K1"] = "创建日期(如果是重复报销，该字段为第一次录入的时间)"
        self.ws["L1"] = "重复报销"
        for i in range(len(invoiceList)):
            self.ws["A" + str(i + 2)] = invoiceList[i].id
            self.ws["B" + str(i + 2)] = invoiceList[i].fileName
            self.ws["C" + str(i + 2)] = invoiceList[i].formName
            self.ws["D" + str(i + 2)] = invoiceList[i].formID
            self.ws["E" + str(i + 2)] = invoiceList[i].toName
            self.ws["F" + str(i + 2)] = invoiceList[i].toID
            self.ws["G" + str(i + 2)] = invoiceList[i].typeName
            self.ws["H" + str(i + 2)] = invoiceList[i].typeID
            self.ws["I" + str(i + 2)] = invoiceList[i].sumPrice
            self.ws["J" + str(i + 2)] = invoiceList[i].invoiceDate

            info = SQLops().find_by_id(invoiceList[i])
            if info is None:
                try:
                    SQLops().add(invoiceList[i])
                except Exception as e:
                    print(e)
                invoiceList[i].repeat = "否"
            else:
                invoiceList[i].createDate = info[10]
                invoiceList[i].repeat = "是"
            self.ws["K" + str(i + 2)] = invoiceList[i].createDate
            self.ws["L" + str(i + 2)] = invoiceList[i].repeat

        self.wb.save(self.filename)


if __name__ == '__main__':
    date1 = str(date.today())
    print(date1)
